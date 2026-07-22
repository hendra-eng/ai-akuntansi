"""
akuntansi_ai.py
================
Modul inti (murni Python, tidak bergantung Streamlit) untuk:
1. Membaca file Excel rekening koran multi-bank (multi-sheet), format kolom
   bebas beda-beda per bank -- kolom dideteksi otomatis berdasarkan nama header,
   bukan posisi tetap per bank.
2. Membaca sheet "COA" (Chart of Accounts) perusahaan bila ada.
3. Mempelajari pola historis: baris yang KOLOM JURNALNYA SUDAH DIISI (NO AKUN /
   NAMA AKUN debet & kredit) dipakai sebagai "contoh yang sudah benar" untuk
   menyusun aturan otomatis (signature keterangan -> pasangan akun).
4. Menerapkan pola tsb + fallback kata kunci COA + fallback AI (Claude) untuk
   baris yang jurnalnya BELUM diisi (data mentah bulan berjalan).

Didesain supaya bisa diuji tanpa perlu menjalankan Streamlit sama sekali.
"""

from __future__ import annotations

import json
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
import openpyxl


# ============================================================
# 1. MEMBACA SHEET COA (Chart of Accounts)
# ============================================================

def muat_coa(wb: openpyxl.Workbook, nama_sheet_coa: str = "COA") -> pd.DataFrame:
    """
    Mengembalikan DataFrame kolom: no_akun, nama_akun, kategori
    Mencari sheet yang namanya mengandung 'coa' (case-insensitive).
    Kalau tidak ada, kembalikan DataFrame kosong (fitur COA asli dinonaktifkan,
    fallback ke kata kunci generik).
    """
    target = None
    for name in wb.sheetnames:
        if "coa" in name.strip().lower():
            target = name
            break
    if target is None:
        return pd.DataFrame(columns=["no_akun", "nama_akun", "kategori"])

    ws = wb[target]
    baris = []
    header_ditemukan = False
    kolom_map = {}
    for row in ws.iter_rows(values_only=True):
        sel = [str(c).strip().lower() if c is not None else "" for c in row]
        if not header_ditemukan:
            if "description" in sel or "nama" in " ".join(sel):
                header_ditemukan = True
                for i, h in enumerate(sel):
                    if h in ("cat", "kategori", "category"):
                        kolom_map["kategori"] = i
                    elif h in ("description", "nama", "nama akun", "namaakun"):
                        kolom_map["nama_akun"] = i
                    elif h == "" and i not in kolom_map.values():
                        # kolom kode akun biasanya tanpa header jelas (Unnamed)
                        kolom_map.setdefault("no_akun", i)
            continue
        if "no_akun" not in kolom_map or "nama_akun" not in kolom_map:
            continue
        no_akun = row[kolom_map["no_akun"]] if kolom_map.get("no_akun") is not None else None
        nama_akun = row[kolom_map["nama_akun"]] if kolom_map.get("nama_akun") is not None else None
        kategori = row[kolom_map["kategori"]] if kolom_map.get("kategori") is not None else None
        if no_akun is None or nama_akun is None:
            continue
        baris.append({"no_akun": no_akun, "nama_akun": str(nama_akun).strip(), "kategori": kategori})

    return pd.DataFrame(baris)


# ============================================================
# 2. MEMBACA SHEET REKENING KORAN (deteksi kolom otomatis)
# ============================================================

def _cari_header_row(ws, max_scan: int = 6):
    """Cari baris header: baris yang mengandung 'keterangan'/'remarks' DAN 'saldo'/'balance'."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        if ("keterangan" in teks or "remarks" in teks) and ("saldo" in teks or "balance" in teks):
            return i + 1, list(row)  # nomor baris 1-indexed di sheet
    return None, None


def _cari_idx(headers, keywords, sampai=None):
    rentang = headers[:sampai] if sampai is not None else headers
    for i, h in enumerate(rentang):
        if h is None:
            continue
        h_low = str(h).lower()
        if any(k in h_low for k in keywords):
            return i
    return None


class FormatTidakDikenali(ValueError):
    pass


def parse_sheet_bank(ws, nama_bank: str) -> pd.DataFrame:
    """
    Parse satu sheet rekening koran menjadi DataFrame kolom standar:
    bank, tanggal, keterangan, mutasi_debet, mutasi_kredit, saldo,
    supplier_cust, voucher,
    no_akun_debet, nama_akun_debet, jml_debet,
    no_akun_kredit, nama_akun_kredit, jml_kredit
    (kolom jurnal berisi NaN kalau memang belum diisi di file sumber -- itu
    artinya sheet ini data mentah yang perlu dikategorikan.)
    """
    header_rownum, header_row = _cari_header_row(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_bank}' tidak dikenali sebagai rekening koran "
            "(tidak ditemukan kolom KETERANGAN/REMARKS + SALDO/BALANCE)."
        )

    headers = header_row
    saldo_idx = _cari_idx(headers, ["saldo", "balance"])
    if saldo_idx is None:
        raise FormatTidakDikenali(f"Kolom SALDO tidak ditemukan di sheet '{nama_bank}'.")

    idx_tanggal = _cari_idx(headers, ["tgl", "tanggal", "date"], sampai=saldo_idx + 1)
    idx_keterangan = _cari_idx(headers, ["keterangan", "remarks"], sampai=saldo_idx + 1)
    idx_debit_stmt = _cari_idx(headers, ["debit", "debet"], sampai=saldo_idx)
    idx_kredit_stmt = _cari_idx(headers, ["kredit", "credit"], sampai=saldo_idx)

    if idx_tanggal is None or idx_keterangan is None or idx_debit_stmt is None or idx_kredit_stmt is None:
        raise FormatTidakDikenali(
            f"Kolom wajib (tanggal/keterangan/debit/kredit) tidak lengkap terdeteksi di sheet '{nama_bank}'."
        )

    idx_supplier = _cari_idx(headers, ["supplier", "cust"])
    idx_voucher = _cari_idx(headers, ["voucher"])

    # Kolom jurnal: dicari via exact match "DEBET"/"KREDIT" (huruf besar semua),
    # ini konsisten membedakan dari kolom statement bank (mis. "Debit Amount",
    # "MUTASI_DEBET", "Kredit") yang penulisannya selalu beda.
    idx_jurnal_debet = None
    idx_jurnal_kredit = None
    for i, h in enumerate(headers):
        if h == "DEBET":
            idx_jurnal_debet = i
        if h == "KREDIT":
            idx_jurnal_kredit = i

    ada_jurnal = idx_jurnal_debet is not None and idx_jurnal_kredit is not None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        if len(row) <= max(idx_tanggal, idx_keterangan):
            continue
        if row[idx_keterangan] is None and row[idx_tanggal] is None:
            continue
        rows.append({
            "bank": nama_bank,
            "tanggal": row[idx_tanggal],
            "keterangan": row[idx_keterangan],
            "mutasi_debet": row[idx_debit_stmt] or 0,
            "mutasi_kredit": row[idx_kredit_stmt] or 0,
            "saldo": row[saldo_idx],
            "supplier_cust": row[idx_supplier] if idx_supplier is not None else None,
            "voucher": row[idx_voucher] if idx_voucher is not None else None,
            "no_akun_debet": row[idx_jurnal_debet - 2] if ada_jurnal else None,
            "nama_akun_debet": row[idx_jurnal_debet - 1] if ada_jurnal else None,
            "jml_debet": row[idx_jurnal_debet] if ada_jurnal else None,
            "no_akun_kredit": row[idx_jurnal_kredit - 2] if ada_jurnal else None,
            "nama_akun_kredit": row[idx_jurnal_kredit - 1] if ada_jurnal else None,
            "jml_kredit": row[idx_jurnal_kredit] if ada_jurnal else None,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce", dayfirst=True)
    return df


def muat_workbook_rekening_koran(file_like) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """
    Baca seluruh workbook. Sheet bernama mengandung 'coa' dipakai sebagai COA.
    Sheet lain diperlakukan sebagai sheet rekening koran per bank (nama sheet
    = nama bank). Sheet yang gagal dikenali formatnya dilewati & dilaporkan.

    Return: (df_gabungan, df_coa, daftar_peringatan)
    """
    wb = openpyxl.load_workbook(file_like, read_only=True, data_only=True)
    df_coa = muat_coa(wb)

    semua_df = []
    peringatan = []
    for nama in wb.sheetnames:
        if "coa" in nama.strip().lower():
            continue
        ws = wb[nama]
        try:
            df = parse_sheet_bank(ws, nama)
            if df.empty:
                peringatan.append(f"Sheet '{nama}' tidak berisi baris data, dilewati.")
                continue
            semua_df.append(df)
        except FormatTidakDikenali as e:
            peringatan.append(str(e))

    if not semua_df:
        df_gabungan = pd.DataFrame()
    else:
        df_gabungan = pd.concat(semua_df, ignore_index=True)

    return df_gabungan, df_coa, peringatan


# ============================================================
# 3. EKSTRAKSI "SIGNATURE" DARI KETERANGAN
# ============================================================

def ekstrak_signature(keterangan) -> str:
    """
    Ambil token pertama dari keterangan sebagai penanda pola (mis. 'NBMB',
    'BRIVA', 'PLNPOST', 'TLKM', 'BPJS', 'ONUS', dst). Angka di ekor token
    dibuang supaya varian nomor tidak memecah pola yang sama.
    """
    if keterangan is None or (isinstance(keterangan, float) and pd.isna(keterangan)):
        return "TIDAK_ADA_KETERANGAN"
    t = str(keterangan).upper().strip()
    if not t:
        return "TIDAK_ADA_KETERANGAN"
    token = t.split(" ")[0].split("/")[0]
    token = re.sub(r"[0-9]+$", "", token)
    return token if token else "TIDAK_ADA_KETERANGAN"


def _arah(row) -> str:
    return "MASUK" if (row.get("mutasi_kredit") or 0) > 0 else "KELUAR"


# ============================================================
# 4. MEMPELAJARI POLA DARI DATA HISTORIS (yang jurnalnya sudah diisi)
# ============================================================

@dataclass
class Pola:
    # key: (signature, arah) -> info
    aturan: dict = field(default_factory=dict)

    def to_dict(self):
        out = {}
        for k, v in self.aturan.items():
            out["||".join(k)] = v
        return out

    @classmethod
    def from_dict(cls, d):
        aturan = {}
        for k, v in d.items():
            sig, arah = k.split("||")
            aturan[(sig, arah)] = v
        return cls(aturan=aturan)


def pelajari_pola(df: pd.DataFrame) -> Pola:
    """
    df harus sudah punya kolom no_akun_debet / no_akun_kredit (baris tanpa
    jurnal otomatis diabaikan). Untuk tiap (signature, arah), hitung pasangan
    akun (debet, kredit) yang paling sering muncul -- juga simpan apakah
    pola tsb 100% konsisten atau cuma mayoritas (perlu direview manual).

    Tambahan: kalau satu signature ternyata punya >1 pasangan akun, coba
    pecah berdasarkan "kelas nominal" (kecil vs besar) -- pola umum di
    rekening koran: nominal kecil (< 20 ribu) sering kali biaya admin bank,
    sedangkan nominal besar adalah transaksi utamanya.
    """
    df_valid = df[df["no_akun_debet"].notna() & df["no_akun_kredit"].notna()].copy()
    if df_valid.empty:
        return Pola()

    df_valid["signature"] = df_valid["keterangan"].apply(ekstrak_signature)
    df_valid["arah"] = df_valid.apply(_arah, axis=1)
    df_valid["nominal"] = df_valid[["mutasi_debet", "mutasi_kredit"]].max(axis=1)

    aturan = {}
    for (sig, arah), g in df_valid.groupby(["signature", "arah"]):
        pasangan = list(zip(g["no_akun_debet"], g["nama_akun_debet"], g["no_akun_kredit"], g["nama_akun_kredit"]))
        counter = Counter(pasangan)
        if len(counter) == 1:
            (nd, nnd, nk, nnk), jumlah = counter.most_common(1)[0]
            aturan[(sig, arah)] = {
                "no_akun_debet": nd, "nama_akun_debet": nnd,
                "no_akun_kredit": nk, "nama_akun_kredit": nnk,
                "konsisten": True, "jumlah_contoh": int(jumlah),
            }
        else:
            # coba pisah kecil vs besar berdasarkan median nominal
            median_nom = g["nominal"].median()
            g_kecil = g[g["nominal"] <= median_nom * 0.2] if median_nom > 0 else g.iloc[0:0]
            g_besar = g[g["nominal"] > median_nom * 0.2] if median_nom > 0 else g

            def _pasangan_dominan(sub):
                if sub.empty:
                    return None
                p = list(zip(sub["no_akun_debet"], sub["nama_akun_debet"], sub["no_akun_kredit"], sub["nama_akun_kredit"]))
                return Counter(p).most_common(1)[0]

            dom_kecil = _pasangan_dominan(g_kecil)
            dom_besar = _pasangan_dominan(g_besar)

            if dom_kecil and dom_besar and dom_kecil[0] != dom_besar[0]:
                (nd, nnd, nk, nnk), j1 = dom_besar
                aturan[(sig, arah)] = {
                    "no_akun_debet": nd, "nama_akun_debet": nnd,
                    "no_akun_kredit": nk, "nama_akun_kredit": nnk,
                    "konsisten": False, "jumlah_contoh": int(j1),
                    "catatan": "Ada varian nominal kecil (kemungkinan biaya admin) - lihat 'pola_nominal_kecil'",
                    "pola_nominal_kecil": {
                        "no_akun_debet": dom_kecil[0][0], "nama_akun_debet": dom_kecil[0][1],
                        "no_akun_kredit": dom_kecil[0][2], "nama_akun_kredit": dom_kecil[0][3],
                        "ambang_nominal": float(median_nom * 0.2),
                    },
                }
            else:
                (nd, nnd, nk, nnk), jumlah = counter.most_common(1)[0]
                aturan[(sig, arah)] = {
                    "no_akun_debet": nd, "nama_akun_debet": nnd,
                    "no_akun_kredit": nk, "nama_akun_kredit": nnk,
                    "konsisten": False, "jumlah_contoh": int(jumlah),
                    "catatan": f"Mayoritas dari {sum(counter.values())} contoh, ada variasi lain - perlu direview.",
                }

    return Pola(aturan=aturan)


def gabung_pola(pola_lama: Pola, pola_baru: Pola) -> Pola:
    """Gabungkan pola baru ke pola lama (pola baru menang kalau ada bentrok signature+arah)."""
    hasil = dict(pola_lama.aturan)
    hasil.update(pola_baru.aturan)
    return Pola(aturan=hasil)


def simpan_pola(pola: Pola, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(pola.to_dict(), f, ensure_ascii=False, indent=2, default=str)


def muat_pola(path: str) -> Pola:
    if not os.path.exists(path):
        return Pola()
    with open(path, "r", encoding="utf-8") as f:
        return Pola.from_dict(json.load(f))


# ============================================================
# 5. FALLBACK KATA KUNCI (kalau pola historis tidak ada match)
# ============================================================

# Peta kata kunci umum -> kemungkinan nama akun (dicocokkan dengan isi COA asli
# perusahaan via 'contains'; kalau tidak ada satupun yang cocok, dipakai
# sebagai label kategori generik langsung).
KATA_KUNCI_AKUN = {
    "listrik": "LISTRIK", "pln": "LISTRIK",
    "telepon": "TELEPON", "telkom": "TELEPON", "internet": "TELEPON", "wifi": "TELEPON",
    "gaji": "GAJI", "payroll": "GAJI",
    "bpjs": "JAMSOSTEK", "jamsostek": "JAMSOSTEK",
    "sewa": "SEWA", "rental": "SEWA", "kontrak": "SEWA",
    "atk": "PERLENGKAPAN", "alat tulis": "PERLENGKAPAN",
    "bensin": "BBM", "pertamina": "BBM", "shell": "BBM", "bbm": "BBM",
    "pajak": "PAJAK", "pph": "PAJAK", "ppn": "PAJAK",
    "admin bank": "ADM BANK", "biaya adm": "ADM BANK", "provisi": "ADM BANK",
    "asuransi": "ASURANSI",
    "promosi": "PROMOSI", "iklan": "PROMOSI",
    "service": "PEMELIHARAAN", "maintenance": "PEMELIHARAAN", "perbaikan": "PEMELIHARAAN",
}


def cocokkan_kata_kunci_ke_coa(keterangan: str, df_coa: pd.DataFrame):
    """Coba cocokkan keterangan ke salah satu nama akun COA asli via kata kunci."""
    if df_coa is None or df_coa.empty:
        return None
    t = str(keterangan).lower()
    for kw, penanda in KATA_KUNCI_AKUN.items():
        if kw in t:
            cocok = df_coa[df_coa["nama_akun"].str.upper().str.contains(penanda, na=False)]
            if not cocok.empty:
                baris = cocok.iloc[0]
                return int(baris["no_akun"]), baris["nama_akun"]
    return None


# ============================================================
# 6. FALLBACK AI (Claude) UNTUK YANG TIDAK TERDETEKSI POLA/KATA KUNCI
# ============================================================

def ambil_api_key():
    try:
        import streamlit as st
        if "ANTHROPIC_API_KEY" in st.secrets:
            return st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        pass
    return os.environ.get("ANTHROPIC_API_KEY")


def kategorikan_dengan_ai(baris_belum_jelas: list[dict], df_coa: pd.DataFrame, api_key: str):
    """
    baris_belum_jelas: list of dict {"idx":..., "keterangan":..., "arah":..., "nominal":...}
    df_coa: daftar akun asli perusahaan (no_akun, nama_akun, kategori)
    Return: dict {idx: {"no_akun_debet":..., "no_akun_kredit":..., "supplier_cust":...}}
    """
    if not baris_belum_jelas:
        return {}
    try:
        import anthropic
    except ImportError:
        return None

    client = anthropic.Anthropic(api_key=api_key)

    if df_coa is not None and not df_coa.empty:
        daftar_akun_str = "\n".join(f"{int(r.no_akun)} - {r.nama_akun}" for r in df_coa.itertuples())
    else:
        daftar_akun_str = "(Tidak ada COA asli diupload; gunakan nama kategori umum sebagai pengganti no_akun, mis. 'BEBAN LISTRIK')"

    daftar_transaksi_str = "\n".join(
        f"{b['idx']}. [{b['arah']}, Rp{b['nominal']:,.0f}] {b['keterangan']}"
        for b in baris_belum_jelas
    )

    prompt = f"""Kamu adalah asisten akuntansi. Tugasmu: menentukan jurnal (akun yang didebit dan
dikredit) untuk tiap transaksi rekening koran berikut, berdasarkan Chart of Accounts (COA)
perusahaan ini:

{daftar_akun_str}

Aturan umum rekening koran bank:
- Transaksi arah MASUK (uang masuk ke rekening bank): akun BANK di-debit, akun lawan (mis.
  PIUTANG USAHA / PENJUALAN / PENDAPATAN LAIN) di-kredit.
- Transaksi arah KELUAR (uang keluar dari rekening bank): akun beban/hutang/aset di-debit,
  akun BANK di-kredit.
- Kalau keterangan menyebut nama orang/perusahaan yang jelas merupakan pihak lawan transaksi,
  isi juga field supplier_cust dengan nama itu (nama singkat, tanpa kode transaksi).

Daftar transaksi (nomor. [arah, nominal] keterangan):
{daftar_transaksi_str}

Jawab HANYA dalam format JSON array yang valid, tanpa teks tambahan, tanpa markdown code fence,
dengan format:
[{{"nomor": 1, "no_akun_debet": 11200003, "no_akun_kredit": 11300003, "supplier_cust": "nama atau null"}}]
"""

    try:
        response = client.messages.create(
            model="claude-sonnet-5",
            max_tokens=4000,
            messages=[{"role": "user", "content": prompt}],
        )
        teks = response.content[0].text.strip()
        teks = re.sub(r"^```(json)?|```$", "", teks.strip(), flags=re.MULTILINE).strip()
        hasil = json.loads(teks)
    except Exception:
        return None

    mapping = {}
    for item in hasil:
        try:
            idx = int(item.get("nomor"))
        except (TypeError, ValueError):
            continue
        mapping[idx] = item
    return mapping


# ============================================================
# 7. PIPELINE UTAMA: TERAPKAN POLA + FALLBACK KE SATU DATAFRAME
# ============================================================

def proses_dataframe(df: pd.DataFrame, df_coa: pd.DataFrame, pola: Pola,
                      pakai_ai: bool = False, api_key: Optional[str] = None) -> pd.DataFrame:
    """
    Untuk baris yang jurnalnya SUDAH ada (no_akun_debet notna) -> tandai sumber
    'Historis (sudah ada)'.
    Untuk baris yang jurnalnya BELUM ada -> terapkan pola -> kata kunci COA ->
    (opsional) AI -> kalau semua gagal, tandai 'Belum Terkategori'.
    """
    df = df.copy()
    df["sumber_kategori"] = None

    sudah_ada_mask = df["no_akun_debet"].notna() & df["no_akun_kredit"].notna()
    df.loc[sudah_ada_mask, "sumber_kategori"] = "Historis (sudah ada di file)"

    perlu_isi = df[~sudah_ada_mask].index.tolist()

    # --- tahap 1: pola historis ---
    belum_selesai = []
    for idx in perlu_isi:
        row = df.loc[idx]
        sig = ekstrak_signature(row["keterangan"])
        arah = _arah(row)
        aturan = pola.aturan.get((sig, arah))
        if aturan is None:
            belum_selesai.append(idx)
            continue

        nominal = max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0)
        dipakai = aturan
        if not aturan.get("konsisten", True) and "pola_nominal_kecil" in aturan:
            ambang = aturan["pola_nominal_kecil"]["ambang_nominal"]
            if nominal <= ambang:
                dipakai = aturan["pola_nominal_kecil"]

        df.at[idx, "no_akun_debet"] = dipakai["no_akun_debet"]
        df.at[idx, "nama_akun_debet"] = dipakai["nama_akun_debet"]
        df.at[idx, "no_akun_kredit"] = dipakai["no_akun_kredit"]
        df.at[idx, "nama_akun_kredit"] = dipakai["nama_akun_kredit"]
        df.at[idx, "jml_debet"] = nominal
        df.at[idx, "jml_kredit"] = nominal
        df.at[idx, "sumber_kategori"] = (
            "Pola historis (konsisten)" if aturan.get("konsisten", True)
            else "Pola historis (mayoritas - perlu cek)"
        )

    # --- tahap 2: kata kunci ke COA asli ---
    masih_belum = []
    for idx in belum_selesai:
        row = df.loc[idx]
        arah = _arah(row)
        cocok = cocokkan_kata_kunci_ke_coa(row["keterangan"], df_coa)
        nominal = max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0)
        if cocok:
            no_akun, nama_akun = cocok
            bank_no_akun = None
            bank_nama = None
            # cari akun bank dari baris lain di sheet yang sama sbg lawan transaksi
            kandidat_bank = df[(df["bank"] == row["bank"]) & df["no_akun_debet"].notna()]
            if not kandidat_bank.empty:
                if arah == "MASUK":
                    bank_no_akun = kandidat_bank.iloc[0]["no_akun_debet"]
                    bank_nama = kandidat_bank.iloc[0]["nama_akun_debet"]
                else:
                    bank_no_akun = kandidat_bank.iloc[0]["no_akun_kredit"]
                    bank_nama = kandidat_bank.iloc[0]["nama_akun_kredit"]
            if arah == "MASUK":
                df.at[idx, "no_akun_debet"] = bank_no_akun
                df.at[idx, "nama_akun_debet"] = bank_nama
                df.at[idx, "no_akun_kredit"] = no_akun
                df.at[idx, "nama_akun_kredit"] = nama_akun
            else:
                df.at[idx, "no_akun_debet"] = no_akun
                df.at[idx, "nama_akun_debet"] = nama_akun
                df.at[idx, "no_akun_kredit"] = bank_no_akun
                df.at[idx, "nama_akun_kredit"] = bank_nama
            df.at[idx, "jml_debet"] = nominal
            df.at[idx, "jml_kredit"] = nominal
            df.at[idx, "sumber_kategori"] = "Kata kunci COA"
        else:
            masih_belum.append(idx)

    # --- tahap 3: AI ---
    if pakai_ai and masih_belum:
        api_key = api_key or ambil_api_key()
        if api_key:
            batch = []
            for idx in masih_belum:
                row = df.loc[idx]
                nominal = max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0)
                batch.append({
                    "idx": idx, "keterangan": row["keterangan"],
                    "arah": _arah(row), "nominal": nominal,
                })
            mapping = kategorikan_dengan_ai(batch, df_coa, api_key)
            if mapping:
                lookup_nama = dict(zip(df_coa["no_akun"], df_coa["nama_akun"])) if not df_coa.empty else {}
                for idx in masih_belum:
                    item = mapping.get(idx)
                    if not item:
                        continue
                    nd = item.get("no_akun_debet")
                    nk = item.get("no_akun_kredit")
                    nominal = max(df.at[idx, "mutasi_debet"] or 0, df.at[idx, "mutasi_kredit"] or 0)
                    df.at[idx, "no_akun_debet"] = nd
                    df.at[idx, "nama_akun_debet"] = lookup_nama.get(nd, nd)
                    df.at[idx, "no_akun_kredit"] = nk
                    df.at[idx, "nama_akun_kredit"] = lookup_nama.get(nk, nk)
                    df.at[idx, "jml_debet"] = nominal
                    df.at[idx, "jml_kredit"] = nominal
                    if item.get("supplier_cust"):
                        df.at[idx, "supplier_cust"] = item["supplier_cust"]
                    df.at[idx, "sumber_kategori"] = "AI (Claude)"

    df["sumber_kategori"] = df["sumber_kategori"].fillna("Belum Terkategori - perlu review manual")
    return df